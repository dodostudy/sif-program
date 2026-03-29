#!/usr/bin/env python3
"""
Excel → JSON 변환 스크립트
SIF 프로그램 Excel 파일의 DB시트와 드롭다운참조 시트를 JSON으로 변환
"""

import json
import os
import sys

try:
    import openpyxl
except ImportError:
    print("openpyxl 설치 중...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl

# 경로 설정
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
SOURCE_FILE = os.path.join(PROJECT_DIR, "source", "260302_SIF 프로그램_공정조회 차트_r2.xlsx")
DATA_DIR = os.path.join(PROJECT_DIR, "data")

DB_JSON = os.path.join(DATA_DIR, "db.json")
DROPDOWN_JSON = os.path.join(DATA_DIR, "dropdown-ref.json")

# DB 시트 컬럼 매핑 (원본 컬럼명 → JSON 키)
COLUMN_MAP = {
    "순": "id",
    "공종": "공종",
    "작업명": "작업명",
    "단위작업명": "단위작업명",
    "KOEN 공정": "KOEN공정",
    "기인물 분류(VlookUp)": "기인물분류",
    "기인물": "기인물",
    "구분(12대 기인물 선택)": "12대기인물구분",
    "위험도 순위 (~12위)": "위험도순위",
    "3년간 사고비중(19~21년)": "3년간사고비중",
    "혹서기": "혹서기",
    "재해형태": "재해형태",
    "재해개요 : 재해시기 추출": "재해개요",
    "재해유발요인  재해형태 추출(떨어짐, 추락 등 >> 맨오른쪽 2글자)": "재해유발요인",
    "재해유발요인 재해형태 추출(떨어짐, 추락 등 >> 맨오른쪽 2글자)": "재해유발요인",
    "위험성 감소대책(예시)": "위험성감소대책",
}


def convert_db_sheet(wb):
    """DB 시트를 JSON 배열로 변환"""
    ws = wb["DB"]

    # 헤더 읽기
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    print(f"  DB 시트 헤더: {headers}")

    # 컬럼 인덱스 매핑
    col_indices = {}
    for i, header in enumerate(headers):
        if header in COLUMN_MAP:
            col_indices[COLUMN_MAP[header]] = i

    print(f"  매핑된 컬럼: {list(col_indices.keys())}")

    records = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        # 빈 행 건너뛰기
        if row[0] is None:
            continue

        record = {"id": row_idx}

        for json_key, col_idx in col_indices.items():
            value = row[col_idx] if col_idx < len(row) else None

            # 타입 변환
            if json_key == "KOEN공정":
                value = str(value).strip().upper() == "YES" if value else False
            elif json_key == "12대기인물구분":
                record["12대기인물"] = str(value).strip() == "12대 기인물" if value else False
                continue
            elif json_key == "위험도순위":
                if value and str(value).strip() != "해당없음":
                    try:
                        value = int(value)
                    except (ValueError, TypeError):
                        value = None
                else:
                    value = None
            elif json_key == "id":
                continue  # 이미 row_idx로 설정
            else:
                value = str(value).strip() if value else ""

            record[json_key] = value

        records.append(record)

    return records


def convert_dropdown_sheet(wb):
    """드롭다운참조 시트에서 계층구조와 기인물분류 매핑 추출"""
    ws = wb["_드롭다운참조"]

    # 전체 데이터를 2D 배열로 읽기
    all_data = []
    for row in ws.iter_rows(values_only=True):
        all_data.append([cell for cell in row])

    if not all_data:
        print("  경고: 드롭다운참조 시트가 비어있습니다")
        return {}

    headers = all_data[0]
    print(f"  드롭다운참조 시트: {len(all_data)}행 x {len(headers)}열")

    # 헤더에서 주요 섹션 찾기
    header_map = {}
    for i, h in enumerate(headers):
        if h:
            header_map[str(h).strip()] = i

    result = {
        "hierarchy": {},       # 공종 → 작업명 → 단위작업명
        "기인물분류": {},       # 분류명 → [기인물 목록]
        "12대기인물": [],       # 12대 기인물 목록
        "공종목록": [],         # 공종 이름 목록
        "기인물목록": [],       # 전체 기인물 목록
    }

    # ====== DB 시트에서 계층구조 직접 추출 (더 안정적) ======
    ws_db = wb["DB"]
    hierarchy = {}
    공종set = set()
    기인물set = set()

    for row in ws_db.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue

        공종 = str(row[1]).strip() if row[1] else ""
        작업명 = str(row[2]).strip() if row[2] else ""
        단위작업명 = str(row[3]).strip() if row[3] else ""
        기인물 = str(row[6]).strip() if row[6] else ""
        기인물분류 = str(row[5]).strip() if row[5] else ""
        구분 = str(row[7]).strip() if row[7] else ""

        if 공종:
            공종set.add(공종)
            if 공종 not in hierarchy:
                hierarchy[공종] = {}
            if 작업명:
                if 작업명 not in hierarchy[공종]:
                    hierarchy[공종][작업명] = set()
                if 단위작업명:
                    hierarchy[공종][작업명].add(단위작업명)

        if 기인물:
            기인물set.add(기인물)

        # 기인물분류 매핑
        if 기인물분류 and 기인물:
            if 기인물분류 not in result["기인물분류"]:
                result["기인물분류"][기인물분류] = set()
            result["기인물분류"][기인물분류].add(기인물)

        # 12대 기인물
        if 구분 == "12대 기인물" and 기인물:
            if 기인물 not in result["12대기인물"]:
                result["12대기인물"].append(기인물)

    # set → sorted list 변환
    def sort_key(name):
        """숫자 접두사로 정렬"""
        import re
        match = re.match(r'(\d+)', name)
        return (int(match.group(1)) if match else 999, name)

    for 공종 in hierarchy:
        for 작업명 in hierarchy[공종]:
            hierarchy[공종][작업명] = sorted(hierarchy[공종][작업명], key=sort_key)
        hierarchy[공종] = dict(sorted(hierarchy[공종].items(), key=lambda x: sort_key(x[0])))

    result["hierarchy"] = dict(sorted(hierarchy.items(), key=lambda x: sort_key(x[0])))
    result["공종목록"] = sorted(공종set, key=sort_key)
    result["기인물목록"] = sorted(기인물set)

    # 기인물분류의 set → list
    for k in result["기인물분류"]:
        result["기인물분류"][k] = sorted(result["기인물분류"][k])

    return result


def main():
    print(f"소스 파일: {SOURCE_FILE}")
    print(f"출력 디렉토리: {DATA_DIR}")

    if not os.path.exists(SOURCE_FILE):
        print(f"오류: 소스 파일을 찾을 수 없습니다: {SOURCE_FILE}")
        sys.exit(1)

    os.makedirs(DATA_DIR, exist_ok=True)

    print("\nExcel 파일 로딩...")
    wb = openpyxl.load_workbook(SOURCE_FILE, data_only=True)
    print(f"  시트 목록: {wb.sheetnames}")

    # 1. DB 시트 변환
    print("\n[1/2] DB 시트 변환 중...")
    records = convert_db_sheet(wb)
    print(f"  총 {len(records)}건 변환 완료")

    # 검증
    공종_unique = set(r["공종"] for r in records if r.get("공종"))
    기인물_unique = set(r["기인물"] for r in records if r.get("기인물"))
    재해형태_unique = set(r["재해형태"] for r in records if r.get("재해형태"))
    print(f"  공종: {len(공종_unique)}개, 기인물: {len(기인물_unique)}개, 재해형태: {len(재해형태_unique)}개")

    with open(DB_JSON, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=None)
    print(f"  저장: {DB_JSON} ({os.path.getsize(DB_JSON) / 1024:.0f}KB)")

    # 2. 드롭다운참조 변환
    print("\n[2/2] 드롭다운참조 변환 중...")
    dropdown = convert_dropdown_sheet(wb)
    print(f"  공종: {len(dropdown['공종목록'])}개")
    print(f"  기인물분류: {len(dropdown['기인물분류'])}개 카테고리")
    print(f"  12대기인물: {len(dropdown['12대기인물'])}개")
    print(f"  계층구조 공종수: {len(dropdown['hierarchy'])}개")

    with open(DROPDOWN_JSON, "w", encoding="utf-8") as f:
        json.dump(dropdown, f, ensure_ascii=False, indent=2)
    print(f"  저장: {DROPDOWN_JSON} ({os.path.getsize(DROPDOWN_JSON) / 1024:.0f}KB)")

    wb.close()
    print("\n변환 완료!")


if __name__ == "__main__":
    main()
